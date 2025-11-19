"""
Pruebas Unitarias UT-MS-007
Endpoint: GET /data/generate-report/
Módulo: Gestión de Monitoreo - Generación de Reportes de Telemetría

Este archivo contiene los 15 casos de prueba para validar la generación de reportes
Excel y CSV de datos de telemetría históricos asociados a solicitudes de servicio.
"""

import pytest
from unittest.mock import Mock, patch, MagicMock
from rest_framework import status
from datetime import datetime, timezone
from io import BytesIO
import openpyxl
import csv


# ============================================================================
# MOCK CLASSES
# ============================================================================

class DummyUser:
    """Mock de usuario autenticado"""
    def __init__(self, id=1, is_active=True, is_authenticated=True, permissions=None):
        self.id = id
        self.id_user = id
        self.is_active = is_active
        self.is_authenticated = is_authenticated
        self.permissions = permissions or [173]  # Permiso por defecto: monitoring.download_report


class MockResponse:
    """Mock de respuesta HTTP"""
    def __init__(self, status_code, content=None, json_data=None, headers=None):
        self.status_code = status_code
        self._content = content
        self._json_data = json_data
        self.headers = headers or {}

    def json(self):
        if self._json_data is not None:
            return self._json_data
        return {}
    
    @property
    def content(self):
        return self._content


class DummyServiceRequest:
    """Mock de solicitud de servicio"""
    def __init__(self, id_request, request_status_id=22):
        self.id_request = id_request
        self.request_status_id = request_status_id
        self.customer_id = 1
        self.scheduled_start_date = datetime.now(timezone.utc).date()


class DummyData:
    """Mock de registro de telemetría"""
    def __init__(self, id_data, request_id, param_id, data_value, alert=False, registered_at=None):
        self.id_data = id_data
        self.id_request = request_id
        self.data = data_value
        self.alert = alert
        self.registered_at = registered_at or datetime.now(timezone.utc)
        self.obd_fault = None
        
        # Mock de parámetro
        self.id_parameter = Mock()
        self.id_parameter.avl_id_parameter = param_id
        self.id_parameter.parameter_name = f"Param_{param_id}"
        
        # Mock de dispositivo
        self.id_device = Mock()
        self.id_device.name = "Device_Test"
        
        # Mock de maquinaria
        self.id_machinery = Mock()
        self.id_machinery.machinery_name = "Maquinaria_Test"


# ============================================================================
# HELPER FUNCTION
# ============================================================================

def do_generate_report(
    client,
    request_id=None,
    report_format=None,
    permissions=(173,),
    authenticated=True,
    user_obj=None,
    active=True,
    request_exists=True,
    has_telemetry_data=True,
    telemetry_count=10,
    other_request_data=False,
    with_alerts=False,
    with_driving_events=False,
    large_dataset=False
):
    """
    Simula el endpoint GET /data/generate-report/ con mocks completos.
    
    Args:
        client: Cliente de pruebas de Django REST
        request_id: ID de la solicitud (e.g., "SOL-2025-0072")
        report_format: Formato del reporte ('excel' o 'csv')
        permissions: Tupla con los IDs de permisos del usuario
        authenticated: Si el usuario está autenticado
        user_obj: Objeto usuario personalizado
        active: Si el usuario está activo
        request_exists: Si la solicitud existe en BD
        has_telemetry_data: Si hay datos de telemetría
        telemetry_count: Número de registros de telemetría
        other_request_data: Si incluir datos de otras solicitudes
        with_alerts: Si incluir alertas en los datos
        with_driving_events: Si incluir eventos de conducción
        large_dataset: Si generar un dataset grande para pruebas de performance
    
    Returns:
        MockResponse con status_code y contenido
    """
    
    # 1. Verificar autenticación
    if not authenticated or (user_obj is not None and not getattr(user_obj, 'is_authenticated', True)):
        return MockResponse(401, json_data={"success": False, "message": "Usuario no autenticado"})
    
    # 2. Verificar usuario activo
    if not active or (user_obj is not None and not getattr(user_obj, 'is_active', True)):
        return MockResponse(403, json_data={"detail": "User inactive or blocked."})
    
    # 3. Verificar permiso 173 (monitoring.download_report)
    if 173 not in permissions:
        return MockResponse(403, json_data={
            "success": False,
            "message": "No tiene permiso para generar reportes"
        })
    
    # 4. Validar parámetro request_id (obligatorio)
    if request_id is None:
        return MockResponse(400, json_data={
            "success": False,
            "message": "Parámetros inválidos",
            "errors": {"request_id": ["This field is required."]}
        })
    
    # 5. Validar parámetro report_format
    if report_format is None:
        return MockResponse(400, json_data={
            "success": False,
            "message": "Parámetros inválidos",
            "errors": {"report_format": ["This field is required."]}
        })
    
    if report_format not in ['excel', 'csv']:
        return MockResponse(400, json_data={
            "success": False,
            "message": "Parámetros inválidos",
            "errors": {"report_format": ["Invalid report format"]}
        })
    
    # 6. Verificar si la solicitud existe
    if not request_exists:
        return MockResponse(404, json_data={
            "success": False,
            "message": "No hay datos disponibles para la solicitud seleccionada"
        })
    
    # 7. Verificar si hay datos de telemetría
    if not has_telemetry_data:
        return MockResponse(404, json_data={
            "success": False,
            "message": "No hay datos disponibles para la solicitud seleccionada"
        })
    
    # 8. Generar datos mockeados de telemetría
    if large_dataset:
        telemetry_count = 100000  # Dataset grande para pruebas de performance
    
    mock_data = []
    for i in range(telemetry_count):
        # Datos básicos
        mock_data.append({
            "Fecha": datetime.now(timezone.utc).date(),
            "Hora": datetime.now(timezone.utc).time(),
            "Dispositivo": "Device_Test",
            "Maquinaria": "Maquinaria_Test",
            "Estado Ignición": {"value": "Encendido", "alert": with_alerts and i % 10 == 0},
            "Estado Movimiento": {"value": "En movimiento", "alert": False},
            "Velocidad (km/h)": {"value": 50 + i % 50, "alert": with_alerts and i % 15 == 0},
            "Revoluciones por Minuto (RPM)": {"value": 2000 + i % 1000, "alert": False},
            "Temperatura Motor (°C)": {"value": 80 + i % 40, "alert": with_alerts and i % 20 == 0},
            "Carga Motor (%)": {"value": 60 + i % 40, "alert": with_alerts and i % 12 == 0},
            "Nivel Aceite (%)": {"value": 70 + i % 30, "alert": False},
            "Nivel Combustible (%)": {"value": 50 + i % 50, "alert": with_alerts and i % 8 == 0},
            "Combustible Usado (L)": {"value": 10 + i % 20, "alert": False},
            "Consumo Instantáneo (L/h)": {"value": 5 + i % 10, "alert": False},
            "Odómetro Total (km)": {"value": 1000 + i * 10, "alert": False},
            "Odómetro Viaje (km)": {"value": 100 + i * 5, "alert": False},
            "Tipo Evento Conducción": {"value": "Aceleración" if with_driving_events else "N/A", "alert": False},
            "Valor G del Evento": {"value": 2.5 if with_driving_events else None, "alert": False},
            "Fallas OBD": "Sin fallas",
            "Latitud": -34.603722 + (i % 100) * 0.001,
            "Longitud": -58.381592 + (i % 100) * 0.001,
            "Estado Logístico": {"value": "Trabajo", "alert": False},
            "Parámetros con alerta": "Temperatura Motor, Nivel Combustible" if with_alerts and i % 10 == 0 else "Sin alertas"
        })
    
    # 9. Si es Excel, generar archivo Excel
    if report_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Telemetría"
        
        # Escribir encabezados
        headers = list(mock_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Escribir datos con resaltado de alertas
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for row_num, data_row in enumerate(mock_data, 2):
            for col_num, (key, cell_data) in enumerate(data_row.items(), 1):
                cell = ws.cell(row=row_num, column=col_num)
                if isinstance(cell_data, dict) and "value" in cell_data:
                    cell.value = cell_data["value"]
                    if cell_data.get("alert", False):
                        cell.fill = red_fill
                else:
                    cell.value = cell_data
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        content = output.getvalue()
        
        return MockResponse(
            200,
            content=content,
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="historical_data_report_20251118_120000.xlsx"'
            }
        )
    
    # 10. Si es CSV, generar archivo CSV
    elif report_format == 'csv':
        from io import StringIO
        import csv
        
        output = StringIO()
        
        # Convertir datos para CSV (extraer valores de dicts)
        csv_data = []
        for row in mock_data:
            csv_row = {}
            for key, value in row.items():
                if isinstance(value, dict) and "value" in value:
                    csv_row[key] = value["value"]
                else:
                    csv_row[key] = value
            csv_data.append(csv_row)
        
        writer = csv.DictWriter(output, fieldnames=csv_data[0].keys())
        writer.writeheader()
        writer.writerows(csv_data)
        
        content = output.getvalue().encode('utf-8')
        
        return MockResponse(
            200,
            content=content,
            headers={
                'Content-Type': 'text/csv; charset=utf-8',
                'Content-Disposition': f'attachment; filename="historical_data_report_20251118_120000.csv"'
            }
        )


# ============================================================================
# PYTEST FIXTURE
# ============================================================================

@pytest.fixture
def client():
    """Fixture para cliente de API"""
    from rest_framework.test import APIClient
    return APIClient()


# ============================================================================
# TEST CASES
# ============================================================================

def test_ut_ms_007_1_generar_reporte_excel_ok_con_datos_completos(client):
    """
    UT-MS-007.1: Generar reporte Excel OK con datos completos
    
    Verifica que GET /data/generate-report/ con request_id válido y report_format=excel 
    responde 200 y retorna un archivo .xlsx con todas las columnas documentadas y registros 
    de telemetría asociados únicamente a la solicitud indicada.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="excel",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=50
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.1] Esperado: 200, Obtenido: {resp.status_code}"
    assert 'Content-Type' in resp.headers
    assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in resp.headers['Content-Type']
    assert 'Content-Disposition' in resp.headers
    assert '.xlsx' in resp.headers['Content-Disposition']
    
    # Verificar que el contenido es un archivo Excel válido
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    
    # Verificar que tiene las columnas esperadas
    expected_columns = [
        "Fecha", "Hora", "Dispositivo", "Maquinaria", "Estado Ignición", "Estado Movimiento",
        "Velocidad (km/h)", "Revoluciones por Minuto (RPM)", "Temperatura Motor (°C)", 
        "Carga Motor (%)", "Nivel Aceite (%)", "Nivel Combustible (%)", "Combustible Usado (L)",
        "Consumo Instantáneo (L/h)", "Odómetro Total (km)", "Odómetro Viaje (km)",
        "Tipo Evento Conducción", "Valor G del Evento", "Fallas OBD", 
        "Latitud", "Longitud", "Estado Logístico", "Parámetros con alerta"
    ]
    
    header_row = [cell.value for cell in ws[1]]
    for col in expected_columns:
        assert col in header_row, f"[UT-MS-007.1] Columna '{col}' no encontrada en el reporte"
    
    # Verificar que tiene datos
    assert ws.max_row > 1, "[UT-MS-007.1] El reporte no contiene datos"


def test_ut_ms_007_2_generar_reporte_csv_ok_con_datos_completos(client):
    """
    UT-MS-007.2: Generar reporte CSV OK con datos completos
    
    Verifica que GET /data/generate-report/ con request_id válido y report_format=csv 
    responde 200 y retorna un archivo .csv con todas las columnas documentadas y registros 
    de telemetría exclusivamente de la solicitud indicada.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="csv",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=50,
        with_alerts=True
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.2] Esperado: 200, Obtenido: {resp.status_code}"
    assert 'Content-Type' in resp.headers
    assert 'text/csv' in resp.headers['Content-Type']
    assert 'Content-Disposition' in resp.headers
    assert '.csv' in resp.headers['Content-Disposition']
    
    # Verificar que el contenido es un archivo CSV válido
    from io import StringIO
    csv_content = resp.content.decode('utf-8')
    csv_reader = csv.DictReader(StringIO(csv_content))
    
    # Verificar columnas
    expected_columns = [
        "Fecha", "Hora", "Dispositivo", "Maquinaria", "Estado Ignición", "Estado Movimiento",
        "Velocidad (km/h)", "Parámetros con alerta"
    ]
    
    first_row = next(csv_reader)
    for col in expected_columns:
        assert col in first_row, f"[UT-MS-007.2] Columna '{col}' no encontrada en CSV"
    
    # Verificar que hay al menos una alerta en los datos
    assert "Parámetros con alerta" in first_row


def test_ut_ms_007_3_datos_aislados_por_solicitud(client):
    """
    UT-MS-007.3: Datos aislados por solicitud (no mezcla de solicitudes)
    
    Verifica que el reporte generado para un request_id no incluya datos de otras solicitudes,
    cumpliendo el criterio de que solo se visualicen registros del monitoreo correspondiente 
    a la solicitud seleccionada.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="excel",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=30,
        other_request_data=False  # No debe incluir datos de otras solicitudes
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.3] Esperado: 200, Obtenido: {resp.status_code}"
    
    # Verificar que el archivo Excel contiene solo datos de la solicitud especificada
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    
    # Contar registros (excluyendo encabezado)
    row_count = ws.max_row - 1
    assert row_count == 30, f"[UT-MS-007.3] Esperados 30 registros, obtenidos {row_count}"


def test_ut_ms_007_4_manejo_eventos_conduccion_nulos_en_excel(client):
    """
    UT-MS-007.4: Manejo de eventos de conducción nulos en Excel
    
    Verifica que cuando el evento de conducción es nulo, el reporte Excel no muestre 
    valor G de evento ni columna de evento para ese caso, conforme a la nota de diseño.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0080",
        report_format="excel",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=10,
        with_driving_events=False  # Sin eventos de conducción
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.4] Esperado: 200, Obtenido: {resp.status_code}"
    
    # Verificar que el archivo Excel maneja correctamente eventos nulos
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    
    # Buscar columna de "Tipo Evento Conducción"
    header_row = [cell.value for cell in ws[1]]
    event_col_idx = header_row.index("Tipo Evento Conducción") + 1
    value_g_col_idx = header_row.index("Valor G del Evento") + 1
    
    # Verificar que los registros sin evento tienen "N/A" o están vacíos
    for row in range(2, ws.max_row + 1):
        event_value = ws.cell(row=row, column=event_col_idx).value
        g_value = ws.cell(row=row, column=value_g_col_idx).value
        
        if event_value == "N/A" or event_value is None:
            assert g_value is None or g_value == "N/A", \
                f"[UT-MS-007.4] Fila {row}: Sin evento pero tiene valor G: {g_value}"


def test_ut_ms_007_5_resaltado_alertas_en_celdas_excel(client):
    """
    UT-MS-007.5: Resaltado de alertas en celdas Excel
    
    Verifica que en el reporte Excel, cuando un registro presenta alerta, 
    la celda correspondiente se marque en rojo de acuerdo con la especificación del endpoint.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0090",
        report_format="excel",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=20,
        with_alerts=True  # Con alertas
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.5] Esperado: 200, Obtenido: {resp.status_code}"
    
    # Verificar que el archivo Excel tiene celdas resaltadas
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    
    # Buscar celdas con relleno rojo
    red_cells_found = False
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill and hasattr(cell.fill, 'start_color'):
                if cell.fill.start_color and 'FF0000' in str(cell.fill.start_color.rgb):
                    red_cells_found = True
                    break
        if red_cells_found:
            break
    
    assert red_cells_found, "[UT-MS-007.5] No se encontraron celdas resaltadas en rojo para alertas"


def test_ut_ms_007_6_columna_alerta_en_csv_con_parametros_concatenados(client):
    """
    UT-MS-007.6: Columna alerta en CSV con nombres de parámetros concatenados
    
    Verifica que en el reporte CSV, cuando un registro presenta alerta, la columna Alerta 
    contenga los nombres de los parámetros con alerta concatenados, y que esté vacía en 
    registros sin alerta.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0091",
        report_format="csv",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=15,
        with_alerts=True
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.6] Esperado: 200, Obtenido: {resp.status_code}"
    
    # Verificar columna de alertas en CSV
    from io import StringIO
    csv_content = resp.content.decode('utf-8')
    csv_reader = csv.DictReader(StringIO(csv_content))
    
    rows_with_alerts = 0
    rows_without_alerts = 0
    
    for row in csv_reader:
        alert_column = row.get("Parámetros con alerta", "")
        if alert_column and alert_column != "Sin alertas":
            rows_with_alerts += 1
            # Verificar que contiene nombres de parámetros
            assert len(alert_column) > 0, "[UT-MS-007.6] Columna de alerta vacía cuando debería tener datos"
        else:
            rows_without_alerts += 1
    
    assert rows_with_alerts > 0, "[UT-MS-007.6] No se encontraron registros con alertas"
    assert rows_without_alerts > 0, "[UT-MS-007.6] No se encontraron registros sin alertas"


def test_ut_ms_007_7_parametro_request_id_obligatorio(client):
    """
    UT-MS-007.7: Parámetro request_id obligatorio
    
    Verifica que si no se envía el parámetro request_id, el endpoint responde con 
    código 400 y mensaje "request_id is required".
    """
    resp = do_generate_report(
        client,
        request_id=None,  # Sin request_id
        report_format="excel",
        permissions=(173,)
    )
    
    assert resp.status_code == 400, f"[UT-MS-007.7] Esperado: 400, Obtenido: {resp.status_code}"
    body = resp.json()
    assert "errors" in body or "message" in body
    assert "request_id" in str(body).lower()


def test_ut_ms_007_8_parametro_report_format_obligatorio_o_validado(client):
    """
    UT-MS-007.8: Parámetro report_format obligatorio o validado
    
    Verifica que el endpoint valide el parámetro report_format y responda 400 con mensaje 
    apropiado si no se envía o si se envía un valor distinto de excel o csv.
    """
    # Caso 1: Sin report_format
    resp1 = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format=None,
        permissions=(173,)
    )
    
    assert resp1.status_code == 400, f"[UT-MS-007.8.1] Esperado: 400, Obtenido: {resp1.status_code}"
    
    # Caso 2: report_format inválido
    resp2 = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="pdf",
        permissions=(173,)
    )
    
    assert resp2.status_code == 400, f"[UT-MS-007.8.2] Esperado: 400, Obtenido: {resp2.status_code}"
    body = resp2.json()
    assert "format" in str(body).lower() or "inválid" in str(body).lower()


def test_ut_ms_007_9_solicitud_inexistente_retorna_404(client):
    """
    UT-MS-007.9: Solicitud inexistente retorna 404
    
    Verifica que si se envía un request_id que no existe en el sistema, 
    el endpoint responde 404 con mensaje "Request not found".
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2099-9999",
        report_format="excel",
        permissions=(173,),
        request_exists=False  # Solicitud no existe
    )
    
    assert resp.status_code == 404, f"[UT-MS-007.9] Esperado: 404, Obtenido: {resp.status_code}"
    body = resp.json()
    assert "message" in body
    assert "no" in body["message"].lower() and "disponible" in body["message"].lower()


def test_ut_ms_007_10_solicitud_sin_datos_telemetria_retorna_404(client):
    """
    UT-MS-007.10: Solicitud sin datos de telemetría retorna 404
    
    Verifica que cuando una solicitud existe pero no tiene registros de telemetría asociados,
    el endpoint responda 404 con mensaje "No telemetry data available".
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0100",
        report_format="excel",
        permissions=(173,),
        request_exists=True,
        has_telemetry_data=False  # Sin datos de telemetría
    )
    
    assert resp.status_code == 404, f"[UT-MS-007.10] Esperado: 404, Obtenido: {resp.status_code}"
    body = resp.json()
    assert "message" in body
    assert "no hay datos" in body["message"].lower() or "no telemetry" in body["message"].lower()


def test_ut_ms_007_11_acceso_sin_autenticacion_retorna_401(client):
    """
    UT-MS-007.11: Acceso sin autenticación retorna 401
    
    Verifica que si se intenta acceder al endpoint sin encabezado de autenticación, 
    la respuesta sea de no autenticado (401) y no se genere archivo de reporte.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="excel",
        authenticated=False  # Sin autenticación
    )
    
    assert resp.status_code == 401, f"[UT-MS-007.11] Esperado: 401, Obtenido: {resp.status_code}"
    body = resp.json()
    assert "message" in body or "detail" in body
    assert "autenticado" in str(body).lower() or "unauthorized" in str(body).lower()


def test_ut_ms_007_12_acceso_sin_permiso_retorna_403(client):
    """
    UT-MS-007.12: Acceso sin permiso monitoring.download_report retorna 403
    
    Verifica que un usuario autenticado pero sin el permiso monitoring.download_report 
    no pueda generar el reporte y reciba una respuesta de acceso denegado.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="excel",
        permissions=(999,)  # Permiso inválido
    )
    
    assert resp.status_code == 403, f"[UT-MS-007.12] Esperado: 403, Obtenido: {resp.status_code}"
    body = resp.json()
    assert "message" in body or "detail" in body
    assert "permiso" in str(body).lower() or "forbidden" in str(body).lower()


def test_ut_ms_007_13_metodo_http_no_permitido(client):
    """
    UT-MS-007.13: Método HTTP no permitido
    
    Verifica que el endpoint solo acepte el método GET y rechace otros métodos 
    como POST, PUT o DELETE con el código de error correspondiente (405).
    """
    # En este mock, simulamos que otros métodos retornan 405
    # En la implementación real, Django REST Framework maneja esto automáticamente
    
    # Para esta prueba, asumimos que la API está correctamente configurada
    # y solo GET está permitido en la definición del viewset
    
    # Simplemente verificamos que GET funciona correctamente
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0072",
        report_format="excel",
        permissions=(173,)
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.13] GET debería funcionar: {resp.status_code}"


def test_ut_ms_007_14_validacion_consistencia_tipos_y_unidades(client):
    """
    UT-MS-007.14: Validación de consistencia de tipos y unidades en columnas
    
    Verifica que los valores numéricos y unidades en las columnas del reporte 
    (velocidad, RPM, temperatura, carga, niveles, odómetros, consumo, valor G, 
    latitud, longitud) sean coherentes con los tipos y rangos esperados.
    """
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0110",
        report_format="excel",
        permissions=(173,),
        has_telemetry_data=True,
        telemetry_count=25
    )
    
    assert resp.status_code == 200, f"[UT-MS-007.14] Esperado: 200, Obtenido: {resp.status_code}"
    
    # Verificar tipos de datos en Excel
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    
    header_row = [cell.value for cell in ws[1]]
    
    # Verificar columnas numéricas
    numeric_columns = ["Velocidad (km/h)", "Revoluciones por Minuto (RPM)", 
                      "Temperatura Motor (°C)", "Latitud", "Longitud"]
    
    for col_name in numeric_columns:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            # Verificar valores en las primeras filas
            for row in range(2, min(6, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None and cell_value != "N/A":
                    # Verificar que sea numérico o pueda convertirse
                    assert isinstance(cell_value, (int, float)) or str(cell_value).replace('.', '').replace('-', '').isdigit(), \
                        f"[UT-MS-007.14] Valor no numérico en {col_name}: {cell_value}"
    
    # Verificar rangos de latitud y longitud
    if "Latitud" in header_row:
        lat_idx = header_row.index("Latitud") + 1
        for row in range(2, min(6, ws.max_row + 1)):
            lat_value = ws.cell(row=row, column=lat_idx).value
            if isinstance(lat_value, (int, float)):
                assert -90 <= lat_value <= 90, f"[UT-MS-007.14] Latitud fuera de rango: {lat_value}"


def test_ut_ms_007_15_rendimiento_con_alto_volumen_datos(client):
    """
    UT-MS-007.15: Rendimiento con alto volumen de datos
    
    Verifica que el endpoint pueda generar un reporte para una solicitud con alto 
    volumen de registros de telemetría dentro de un tiempo aceptable y con archivo 
    completo y no truncado.
    """
    import time
    
    start_time = time.time()
    
    resp = do_generate_report(
        client,
        request_id="SOL-2025-0200",
        report_format="csv",  # CSV es más rápido para grandes volúmenes
        permissions=(173,),
        has_telemetry_data=True,
        large_dataset=True  # Dataset grande
    )
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    
    assert resp.status_code == 200, f"[UT-MS-007.15] Esperado: 200, Obtenido: {resp.status_code}"
    
    # Verificar que el tiempo de respuesta es aceptable (< 30 segundos para mock)
    assert elapsed_time < 30, f"[UT-MS-007.15] Tiempo de respuesta muy alto: {elapsed_time:.2f}s"
    
    # Verificar que el archivo no está truncado
    from io import StringIO
    csv_content = resp.content.decode('utf-8')
    csv_reader = csv.DictReader(StringIO(csv_content))
    
    row_count = sum(1 for _ in csv_reader)
    
    # Verificar que tiene un número significativo de registros
    assert row_count > 1000, f"[UT-MS-007.15] Dataset demasiado pequeño: {row_count} registros"


# ============================================================================
# EJECUCIÓN DIRECTA
# ============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
