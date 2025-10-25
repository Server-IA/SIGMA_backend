import csv
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from service_requests.utils.external_user_helper import get_user_display_name


def _format_currency(amount, currency_unit=None):
    """Formatea un monto como moneda con separadores de miles."""
    if amount is None:
        return ""
    
    try:
        # Formatear con separador de miles y 2 decimales
        formatted = f"{float(amount):,.2f}"
        if currency_unit and hasattr(currency_unit, 'symbol'):
            return f"{formatted} {currency_unit.symbol}"
        return formatted
    except (ValueError, TypeError):
        return ""


def _format_date_iso(date_obj):
    """Formatea una fecha en formato ISO YYYY-MM-DD HH:MM:SS."""
    if not date_obj:
        return ""
    
    if hasattr(date_obj, 'date'):
        # Es un datetime
        return date_obj.strftime('%Y-%m-%d %H:%M:%S')
    else:
        # Es una fecha
        return date_obj.strftime('%Y-%m-%d')


def _format_date_only(date_obj):
    """Formatea una fecha solo con fecha (sin hora)."""
    if not date_obj:
        return ""
    
    if hasattr(date_obj, 'date'):
        # Es un datetime, extraer solo la fecha
        return date_obj.date().strftime('%Y-%m-%d')
    else:
        # Es una fecha
        return date_obj.strftime('%Y-%m-%d')


def _format_area_with_unit(area, area_unit):
    """Formatea área con su unidad."""
    if area is None:
        return "N/A"
    
    try:
        area_str = f"{float(area):,.2f}"
        if area_unit and hasattr(area_unit, 'symbol'):
            return f"{area_str} {area_unit.symbol}"
        return area_str
    except (ValueError, TypeError):
        return "N/A"


def _format_altitude_with_unit(altitude, altitude_unit):
    """Formatea altitud con su unidad."""
    if altitude is None:
        return "N/A"
    
    try:
        altitude_str = f"{float(altitude):,.2f}"
        if altitude_unit and hasattr(altitude_unit, 'symbol'):
            return f"{altitude_str} {altitude_unit.symbol}"
        return altitude_str
    except (ValueError, TypeError):
        return "N/A"


def _build_report_data(queryset, user_data_map: Dict[int, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Construye los datos del reporte a partir del queryset y el mapa de usuarios.
    
    Args:
        queryset: QuerySet de ServiceRequest
        user_data_map: Diccionario mapeando user_id -> user_data
        
    Returns:
        Lista de diccionarios con los datos del reporte
    """
    report_data = []
    
    for request in queryset:
        # Datos básicos de la solicitud
        row_data = {
            'codigo_seguimiento': request.id_request,
            'estado_solicitud': request.request_status.name if request.request_status else "",
            'fecha_registro': _format_date_iso(request.creation_date),
            'fecha_programada': _format_date_only(request.scheduled_start_date),
            'fecha_realizacion': _format_date_iso(request.completion_cancellation_datetime) if request.completion_cancellation_datetime else "",
        }
        
        # Datos del cliente
        customer = request.customer
        if customer:
            # Si el cliente tiene id_user, obtener datos del servicio externo
            if customer.id_user_id and customer.id_user_id in user_data_map:
                external_user = user_data_map[customer.id_user_id]
                # Construir nombre completo desde datos externos
                name_parts = []
                name = external_user.get('name', '').strip()
                first_last_name = external_user.get('first_last_name', '').strip()
                second_last_name = external_user.get('second_last_name', '').strip()
                
                if name:
                    name_parts.append(name)
                if first_last_name:
                    name_parts.append(first_last_name)
                if second_last_name:
                    name_parts.append(second_last_name)
                
                customer_name = ' '.join(name_parts)
                customer_document_type = external_user.get('type_document_name', '') or ""
                customer_document = str(external_user.get('document_number', '')) if external_user.get('document_number') else ""
            else:
                # Si no tiene id_user o no hay datos externos, usar datos de la tabla customers
                name_parts = []
                if customer.name:
                    name_parts.append(customer.name)
                if customer.first_last_name:
                    name_parts.append(customer.first_last_name)
                if customer.second_last_name:
                    name_parts.append(customer.second_last_name)
                
                customer_name = ' '.join(name_parts)
                customer_document_type = customer.type_document_id.name if customer.type_document_id else ""
                customer_document = str(customer.document_number) if customer.document_number else ""
            
            row_data.update({
                'cliente_nombre': customer_name,
                'cliente_tipo_documento': customer_document_type,
                'cliente_documento': customer_document,
            })
        else:
            row_data.update({
                'cliente_nombre': "",
                'cliente_tipo_documento': "",
                'cliente_documento': "",
            })
        
        # Maquinarias concatenadas
        machinery_list = []
        for machinery_user in request.machinery_users.all():
            machinery = machinery_user.machinery
            if machinery:
                machinery_name = machinery.machinery_name or ""
                serial_number = machinery.serial_number or ""
                machinery_str = f"{machinery_name} ({serial_number})" if serial_number else machinery_name
                machinery_list.append(machinery_str)
        
        row_data['maquinaria'] = ", ".join(machinery_list)
        
        # Operarios concatenados
        operators_list = []
        for machinery_user in request.machinery_users.all():
            user = machinery_user.user
            if user and user.id_user in user_data_map:
                operator_name = get_user_display_name(user_data_map[user.id_user])
                if operator_name:
                    operators_list.append(operator_name)
        
        row_data['operario'] = ", ".join(operators_list)
        
        # Datos de ubicación
        location = request.request_location
        if location:
            row_data.update({
                'ubic_region': location.department or "",
                'ubic_municipio': str(location.city_id) if location.city_id else "",
                'ubic_lugar': location.place_name or "",
                'ubic_area': _format_area_with_unit(location.area, location.area_unit),
                'ubic_altitud': _format_altitude_with_unit(location.altitude, location.altitude_unit),
            })
        else:
            row_data.update({
                'ubic_region': "",
                'ubic_municipio': "",
                'ubic_lugar': "",
                'ubic_area': "N/A",
                'ubic_altitud': "N/A",
            })
        
        # Datos de pago
        row_data.update({
            'monto_a_pagar': _format_currency(request.amount_to_pay, request.currency_unit_amount_to_pay),
            'cantidad_pagada': _format_currency(request.amount_paid, request.currency_unit_amount_paid),
            'estado_pago': request.payment_status.name if request.payment_status else "",
            'modalidad_pago': request.payment_method.name if request.payment_method else "",
        })
        
        # Observación
        row_data['observacion'] = request.completion_cancellation_observations or ""
        
        report_data.append(row_data)
    
    return report_data


def generate_excel_report(queryset, user_data_map: Dict[int, Dict[str, Any]], user_info: Dict[str, Any] = None) -> bytes:
    """
    Genera un reporte en formato Excel.
    
    Args:
        queryset: QuerySet de ServiceRequest
        user_data_map: Diccionario mapeando user_id -> user_data
        
    Returns:
        Bytes del archivo Excel
    """
    # Construir datos del reporte
    report_data = _build_report_data(queryset, user_data_map)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Solicitudes"
    
    # Agregar fila informativa en la primera fila
    from datetime import datetime
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Información del usuario
    user_name = "Usuario no identificado"
    if user_info:
        user_name = user_info.get('name', 'Usuario no identificado')
    
    # Escribir fila informativa
    info_row = [
        f"Usuario: {user_name}",
        f"Fecha y hora: {current_time}",
        "Formato: Excel"
    ]
    
    # Escribir información en la primera fila (combinar celdas para que se vea bien)
    for col_num, info_text in enumerate(info_row, 1):
        cell = ws.cell(row=1, column=col_num, value=info_text)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Definir columnas (empezar desde la fila 3)
    columns = [
        "Código Seguimiento",
        "Estado Solicitud", 
        "Fecha Registro",
        "Fecha Programada",
        "Fecha Realización",
        "Cliente (Nombre)",
        "Cliente (Tipo Documento)",
        "Cliente (Documento)",
        "Maquinaria",
        "Operario (Nombre)",
        "Ubic. Región",
        "Ubic. Municipio", 
        "Ubic. Lugar",
        "Ubic. Área (m²)",
        "Ubic. Altitud (msnm)",
        "Monto a Pagar",
        "Cantidad Pagada",
        "Estado Pago",
        "Modalidad Pago",
        "Observación"
    ]
    
    # Mapeo de campos a columnas
    field_mapping = [
        'codigo_seguimiento', 'estado_solicitud', 'fecha_registro', 'fecha_programada',
        'fecha_realizacion', 'cliente_nombre', 'cliente_tipo_documento', 'cliente_documento',
        'maquinaria', 'operario', 'ubic_region', 'ubic_municipio', 'ubic_lugar',
        'ubic_area', 'ubic_altitud', 'monto_a_pagar', 'cantidad_pagada', 'estado_pago',
        'modalidad_pago', 'observacion'
    ]
    
    # Estilos
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados (fila 3)
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_num, value=column_title)
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # Escribir datos (empezar desde fila 4)
    for row_num, data_row in enumerate(report_data, 4):
        for col_num, field_name in enumerate(field_mapping, 1):
            value = data_row.get(field_name, "")
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    # Guardar en bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_csv_report(queryset, user_data_map: Dict[int, Dict[str, Any]], user_info: Dict[str, Any] = None) -> str:
    """
    Genera un reporte en formato CSV.
    
    Args:
        queryset: QuerySet de ServiceRequest
        user_data_map: Diccionario mapeando user_id -> user_data
        user_info: Información del usuario que genera el reporte
        
    Returns:
        String del archivo CSV con BOM UTF-8
    """
    # Construir datos del reporte
    report_data = _build_report_data(queryset, user_data_map)
    
    # Definir columnas
    columns = [
        "Código Seguimiento",
        "Estado Solicitud", 
        "Fecha Registro",
        "Fecha Programada",
        "Fecha Realización",
        "Cliente (Nombre)",
        "Cliente (Tipo Documento)",
        "Cliente (Documento)",
        "Maquinaria",
        "Operario (Nombre)",
        "Ubic. Región",
        "Ubic. Municipio", 
        "Ubic. Lugar",
        "Ubic. Área (m²)",
        "Ubic. Altitud (msnm)",
        "Monto a Pagar",
        "Cantidad Pagada",
        "Estado Pago",
        "Modalidad Pago",
        "Observación"
    ]
    
    # Mapeo de campos a columnas
    field_mapping = [
        'codigo_seguimiento', 'estado_solicitud', 'fecha_registro', 'fecha_programada',
        'fecha_realizacion', 'cliente_nombre', 'cliente_tipo_documento', 'cliente_documento',
        'maquinaria', 'operario', 'ubic_region', 'ubic_municipio', 'ubic_lugar',
        'ubic_area', 'ubic_altitud', 'monto_a_pagar', 'cantidad_pagada', 'estado_pago',
        'modalidad_pago', 'observacion'
    ]
    
    # Crear CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Agregar fila informativa en la primera fila
    from datetime import datetime
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Información del usuario
    user_name = "Usuario no identificado"
    if user_info:
        user_name = user_info.get('name', 'Usuario no identificado')
    
    # Escribir fila informativa
    info_row = [
        f"Usuario: {user_name}",
        f"Fecha y hora: {current_time}",
        "Formato: CSV"
    ]
    
    # Escribir información en la primera fila
    writer.writerow(info_row)
    
    # Escribir fila vacía para separación
    writer.writerow([])
    
    # Escribir encabezados
    writer.writerow(columns)
    
    # Escribir datos
    for data_row in report_data:
        row = [data_row.get(field_name, "") for field_name in field_mapping]
        writer.writerow(row)
    
    # Agregar BOM UTF-8 para compatibilidad con Excel
    csv_content = output.getvalue()
    return '\ufeff' + csv_content
