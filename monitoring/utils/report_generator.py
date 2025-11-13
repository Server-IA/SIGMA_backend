from typing import List, Dict, Any
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def _build_report_data(queryset) -> list[Dict[str, Any]]:
    """Construye los datos del reporte a partir del queryset."""
    grouped_data = defaultdict(lambda: {
        "latitudes": [], 
        "longitudes": [], 
        "alerts": {},  
        "values": {}
    })
    

    for record in queryset.select_related('id_parameter', 'id_device', 'id_machinery'):
        key = (record.registered_at, record.id_device)
        param_id = str(record.id_parameter.avl_id_parameter).strip()
        param_name = record.id_parameter.parameter_name

        # Coordenadas
        if param_id == "387":
            if len(grouped_data[key]["latitudes"]) == len(grouped_data[key]["longitudes"]):
                grouped_data[key]["latitudes"].append(record.data)
            else:
                grouped_data[key]["longitudes"].append(record.data)
        else:
            grouped_data[key]["values"][param_id] = record.data
        
        grouped_data[key]["alerts"][param_id] = record.alert
        grouped_data[key]['timestamp'] = record.registered_at
        grouped_data[key]['machinery'] = getattr(record.id_machinery, 'machinery_name', 'Desconocido')
        grouped_data[key]['device'] = getattr(record.id_device, 'name', 'Desconocido')
        grouped_data[key]['obd_fault'] = record.obd_fault

    # --- Construcción de filas ---
    report_data = []

    for (timestamp, id_device), values in grouped_data.items():
        ignition_map = {"0": "Apagado", "1": "Encendido"}
        movement_map = {"0": "Detenido", "1": "En movimiento"}
        event_map = {"1": "Aceleración", "2": "Frenado", "3": "Curva"}
        logistic_map = {"1": "Ida", "2": "Trabajo", "3": "Vuelta"}

        vals = values["values"]
        alerts = values["alerts"]

        def normalize_code(v):
            """Convierte '1.0' → '1' para hacer match en los mapas."""
            try:
                return str(int(float(v)))
            except (ValueError, TypeError):
                return str(v)

        def make_value(param_id, default_value=0, transform_fn=None):
            raw_value = vals.get(param_id, default_value)
            if transform_fn:
                normalized = normalize_code(raw_value)
                display_value = transform_fn(normalized)
            else:
                display_value = raw_value
            
            return {
                "value": display_value,
                "alert": alerts.get(param_id, False)
            }

        ignition_state = make_value("239", "N/A", lambda v: ignition_map.get(v, "N/A"))
        movement_state = make_value("240", "N/A", lambda v: movement_map.get(v, "N/A"))
        logistic_state = make_value("-1", "N/A", lambda v: logistic_map.get(v, "N/A"))

        event_type_code = vals.get("253")
        event_type_display = event_map.get(normalize_code(event_type_code), "N/A") if event_type_code else "N/A"
        event_type = {"value": event_type_display, "alert": alerts.get("253", False)}
        event_value = {"value": vals.get("254") if event_type_code else None, "alert": alerts.get("254", False)}

        latitude = values["latitudes"][0] if values["latitudes"] else 0
        longitude = values["longitudes"][0] if values["longitudes"] else 0

        params_with_alerts = [
            name for param_id, name in [
                ("239", "Estado Ignición"),
                ("240", "Estado Movimiento"),
                ("24", "Velocidad"),
                ("36", "RPM"),
                ("32", "Temperatura Motor"),
                ("31", "Carga Motor"),
                ("1159", "Nivel Aceite"),
                ("48", "Nivel Combustible"),
                ("12", "Combustible Usado"),
                ("60", "Consumo Instantáneo"),
                ("16", "Odómetro Total"),
                ("199", "Odómetro Viaje"),
                ("253", "Tipo Evento"),
                ("-1", "Estado Logístico")
            ] if alerts.get(param_id, False)
        ]

        row_data = {
            "Fecha": values['timestamp'].date(),
            "Hora": values['timestamp'].time(),
            "Dispositivo": values.get("device"),
            "Maquinaria": values.get("machinery"),
            "Estado Ignición": ignition_state,
            "Estado Movimiento": movement_state,
            "Velocidad (km/h)": make_value("24", 0),
            "Revoluciones por Minuto (RPM)": make_value("36", 0),
            "Temperatura Motor (°C)": make_value("32", 0),
            "Carga Motor (%)": make_value("31", 0),
            "Nivel Aceite (%)": make_value("1159", 0),
            "Nivel Combustible (%)": make_value("48", 0),
            "Combustible Usado (L)": make_value("12", 0),
            "Consumo Instantáneo (L/h)": make_value("60", 0),
            "Odómetro Total (km)": make_value("16", 0),
            "Odómetro Viaje (km)": make_value("199", 0),
            "Tipo Evento Conducción": event_type,
            "Valor G del Evento": event_value,
            "Fallas OBD": values.get("obd_fault") or "Sin fallas",
            "Latitud": latitude,
            "Longitud": longitude,
            "Estado Logístico": logistic_state,
            "Parámetros con alerta": ", ".join(params_with_alerts) or "Sin alertas",
        }

        report_data.append(row_data)

    return report_data


def generate_excel_report(queryset) -> bytes:
    """Genera un reporte en formato Excel a partir del queryset."""
    report_data = _build_report_data(queryset)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Telemetría"

    headers = list(report_data[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row_num, data_row in enumerate(report_data, 2):
        for col_num, (key, cell_data) in enumerate(data_row.items(), 1):
            cell = ws.cell(row=row_num, column=col_num)
            if isinstance(cell_data, dict) and "value" in cell_data:
                cell.value = cell_data["value"]
                if cell_data.get("alert", False):
                    cell.fill = red_fill
            else:
                cell.value = cell_data

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
